import streamlit as st
import subprocess
import base64
import pandas as pd
import os
import json
import tempfile
import shutil
import re
import time
from datetime import datetime
from itertools import combinations

from openpyxl import load_workbook
from io import BytesIO

# ----------------------------
# PAGE CONFIG
# ----------------------------
st.set_page_config(
    page_title="MXLab Tool",
    page_icon=r"images\logodefinitivo.png",
    layout="wide"
)

FINAL_OUTPUT_DIR = r"C:\Users\JuanGuerrero2\Desktop\MXTest2024new\testRunner"

# ----------------------------
# MXtest RUNNER
# ----------------------------
MXTEST_ROOT = r"C:\Users\JuanGuerrero2\Desktop\MXTest2024new"
LAUNCH_CMD = "LaunchTestPackage.cmd"
DEFAULT_RESULTS_FOLDER = "ResultsFolder"
FIXED_N = "N"

FC_RESULTS_ROOT = os.path.join(MXTEST_ROOT, "File Comparison Results")

# ----------------------------
# PK inference tuning (FAST / SAFE)
# ----------------------------
SAMPLE_ROWS = 1500
MAX_PK_COLS = 4
UNIQ_TARGET = 0.999
TOP_COLS = 10
MAX_COMBOS_PER_K = 300
MAX_FILES_INFERENCE = 250

MEASURE_TOKENS = {
    "pnl", "p&l", "profit", "loss", "gain", "pl",
    "amount", "amt", "cash", "price", "rate", "spread",
    "quantity", "qty", "notional", "npv", "pv", "value",
    "market", "clean", "dirty"
}

STRONG_ID_TOKENS = {
    "nb",
    "trade", "tradeid",
    "contract", "contractid", "contractnumber",
    "deal", "dealid",
    "order", "orderid",
    "transaction", "transactionid",
    "position", "positionid",
    "booking", "bookingid",
}

_num_re = re.compile(r"^-?\d+(\.\d+)?$")

# ----------------------------
# BACKGROUND + STYLE
# ----------------------------
def add_bg_from_local(image_file):
    if not os.path.exists(image_file):
        return
    with open(image_file, "rb") as image:
        encoded = base64.b64encode(image.read()).decode()

    st.markdown(f"""
    <style>
      .stApp {{
        background-image: linear-gradient(120deg, rgba(9,16,30,.84), rgba(0,0,0,.28)),
                          url("data:image/png;base64,{encoded}");
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
      }}
      header {{visibility: hidden;}}
      #MainMenu {{visibility: hidden;}}
      footer {{visibility: hidden;}}

      html, body, [class*="css"] {{
        font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
      }}

      h1, h2, h3, h4, h5, h6 {{
        color: rgba(255,255,255,.96) !important;
      }}
      .stMarkdown, .stMarkdown p, .stMarkdown li {{
        color: rgba(255,255,255,.92);
      }}

      .block-container {{
        padding-top: 1.4rem;
        padding-bottom: 2.0rem;
        max-width: 1650px;
      }}

      .mxl-title {{
        font-size: 2.2rem;
        font-weight: 900;
        margin: 0 0 .15rem 0;
        color: rgba(255,255,255,.96);
      }}
      .mxl-subtitle {{
        font-size: 1rem;
        color: rgba(255,255,255,.78);
        margin: 0 0 1.0rem 0;
      }}
      .mxl-divider {{
        height: 1px;
        background: linear-gradient(90deg, rgba(255,255,255,.0), rgba(255,255,255,.35), rgba(255,255,255,.0));
        margin: .7rem 0 1.1rem 0;
      }}

      [data-testid="stContainer"][data-testid="stContainer"] > div {{
        background: rgba(255,255,255,.08) !important;
        border: 1px solid rgba(255,255,255,.14) !important;
        box-shadow: 0 10px 30px rgba(0,0,0,.35) !important;
        border-radius: 18px !important;
        padding: 16px 16px 14px 16px !important;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
      }}

      label, .stTextInput label {{
        color: rgba(255,255,255,.88) !important;
        font-weight: 650;
      }}

      [data-testid="stTextInput"] input,
      [data-testid="stTextInput"] textarea {{
        background: rgba(255,255,255,.95) !important;
        border: 1px solid rgba(0,0,0,.12) !important;
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        border-radius: 12px !important;
        padding: 10px 12px !important;
      }}

      [data-testid="stDataFrame"] {{
        background: rgba(255,255,255,.90);
        border: 1px solid rgba(0,0,0,.12);
        border-radius: 14px;
        overflow: hidden;
      }}
      [data-testid="stDataFrame"] * {{
        color: rgba(0,0,0,.85) !important;
      }}

      [data-testid="stButton"] > button {{
        width: 100% !important;
        border-radius: 14px !important;
        border: 1px solid rgba(255,255,255,.18) !important;
        padding: 12px 14px !important;
        font-weight: 900 !important;
        letter-spacing: .2px !important;
        background: linear-gradient(135deg, rgba(46,213,115,.95), rgba(0,0,168,.95)) !important;
        color: #061018 !important;
        box-shadow: 0 12px 30px rgba(0,0,0,.35) !important;
        opacity: 1 !important;
        filter: none !important;
      }}
      [data-testid="stButton"] > button:hover {{
        transform: translateY(-1px) !important;
        filter: brightness(1.03) !important;
        background: linear-gradient(135deg, rgba(46,213,115,1), rgba(0,168,255,1)) !important;
        color: #061018 !important;
      }}

      .author-footer {{
        text-align: center;
        font-size: 13px;
        margin-top: 16px;
        color: rgba(255,255,255,.70);
      }}

      /* Tabs: texto blanco siempre */
      [data-testid="stTabs"] button {{
        color: rgba(255,255,255,.92) !important;
        font-weight: 800 !important;
      }}
      /* Tab activo en rojo */
      [data-testid="stTabs"] [aria-selected="true"] {{
        color: #ff3b3b !important;
        border-bottom: 2px solid #ff3b3b !important;
      }}
      [data-testid="stTabs"] div[role="tablist"] > div {{
        border-bottom: 2px solid rgba(255,255,255,.18) !important;
      }}

      /* Expander: título siempre blanco */
      [data-testid="stExpander"] summary,
      [data-testid="stExpander"] summary * {{
        color: rgba(255,255,255,.92) !important;
        font-weight: 800 !important;
      }}
      /* Code dentro del expander: texto negro */
      [data-testid="stExpander"] pre,
      [data-testid="stExpander"] code {{
        color: #111111 !important;
      }}
      /* Spinner: texto blanco */
      [data-testid="stSpinner"] * {{
        color: rgba(255,255,255,.95) !important;
        font-weight: 800 !important;
      }}
      /* Caption */
      .stCaption, .stCaption * {{
        color: rgba(255,255,255,.80) !important;
      }}

      /* ---- PERSONALIZACIÓN PARA LOS BOTONES DE DESCARGA ---- */
      .custom-download-btn > button {{
        width: 100% !important;
        border-radius: 14px !important;
        border: 1px solid rgba(255,255,255,.18) !important;
        padding: 12px 14px !important;
        font-weight: 900 !important;
        letter-spacing: .2px !important;
        background: linear-gradient(135deg, rgba(46,213,115,.95), rgba(0,0,168,.95)) !important;
        color: #061018 !important;
        box-shadow: 0 12px 30px rgba(0,0,0,.35) !important;
        opacity: 1 !important;
        filter: none !important;
      }}
      .custom-download-btn > button:hover {{
        transform: translateY(-1px) !important;
        filter: brightness(1.03) !important;
        background: linear-gradient(135deg, rgba(46,213,115,1), rgba(0,168,255,1)) !important;
        color: #061018 !important;
      }}
    </style>
    """, unsafe_allow_html=True)

add_bg_from_local("images/background.jpg")

# ----------------------------
# Helper: PNG -> base64 data uri (para mostrar logo en el header)
# ----------------------------
def img_to_data_uri(path: str) -> str:
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:image/png;base64,{b64}"

# ----------------------------
# Helpers
# ----------------------------
def normalize_header(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[()]", "", s)
    s = s.replace("&", "")
    return s.upper()

def normalize_headers_list(fields):
    out = []
    for x in (fields or []):
        nx = normalize_header(x)
        if nx != "":
            out.append(nx)
    return out

def normalize_csv_headers_df(df: pd.DataFrame) -> pd.DataFrame:
    cols = [normalize_header(c) for c in df.columns]
    seen = {}
    fixed = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            fixed.append(c)
        else:
            seen[c] += 1
            fixed.append(f"{c}_{seen[c]}")
    df.columns = fixed
    return df

def normalize_pk_list(pk_list):
    return normalize_headers_list(pk_list)

def detect_delimiter(path):
    candidates = [",", ";", "|", "\t"]
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            line = f.readline()
        counts = {d: line.count(d) for d in candidates}
        return max(counts, key=counts.get)
    except Exception:
        return ","

def _safe_rename_existing(path: str) -> str | None:
    if not os.path.exists(path):
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_path = f"{path}__old__{ts}"
    os.rename(path, new_path)
    return new_path

def move_package_to_final_location(pkg_dir: str, final_root: str) -> str:
    os.makedirs(final_root, exist_ok=True)
    src_path = os.path.abspath(pkg_dir)
    dst_path = os.path.join(final_root, pkg_dir)

    if os.path.exists(dst_path):
        try:
            shutil.rmtree(dst_path)
        except PermissionError:
            _safe_rename_existing(dst_path)

    last_err = None
    for _ in range(5):
        try:
            shutil.move(src_path, dst_path)
            return dst_path
        except PermissionError as e:
            last_err = e
            time.sleep(0.5)

    raise PermissionError(
        f"Could not move package to destination (locked).\n"
        f"Close MXtest/Explorer if they are using it and try again.\n"
        f"Source: {src_path}\nDestination: {dst_path}\nError: {last_err}"
    )

def _norm_name(col: str) -> str:
    return (
        str(col).strip().lower()
        .replace(" ", "")
        .replace("-", "")
        .replace(".", "")
        .replace("/", "")
        .replace("\\", "")
        .replace("_", "")
    )

def looks_like_measure_name(col: str) -> bool:
    n = _norm_name(col)
    return any(tok in n for tok in MEASURE_TOKENS)

def looks_like_id_name(col: str) -> bool:
    n = _norm_name(col)
    return "id" in n

def looks_like_strong_id_name(col: str) -> bool:
    n = _norm_name(col)
    if n == "nb" or n.startswith("nb") or n.endswith("nb") or "nb" in n:
        return True
    if "trade" in n:
        return True
    if "contract" in n:
        return True
    for tok in STRONG_ID_TOKENS:
        if tok != "nb" and tok in n:
            return True
    return False

def numeric_profile(series: pd.Series):
    s = series.astype(str).str.strip()
    s = s[s != ""]
    if len(s) == 0:
        return (0.0, 0.0)
    is_num = s.apply(lambda x: bool(_num_re.match(x)))
    numeric_ratio = float(is_num.mean())
    dec_ratio = float(s[is_num].str.contains(r"\.", regex=True).mean()) if is_num.any() else 0.0
    return (numeric_ratio, dec_ratio)

def read_csv_robust(file_path, sep, nrows, header):
    encodings_to_try = ["utf-8", "utf-8-sig", "utf-16", "utf-16le", "utf-16be", "cp1252", "latin1"]
    last_err = None
    for enc in encodings_to_try:
        try:
            return pd.read_csv(
                file_path,
                sep=sep,
                dtype=str,
                nrows=nrows,
                header=header,
                engine="python",
                encoding=enc,
                encoding_errors="replace"
            )
        except Exception as e:
            last_err = e
            continue
    raise last_err

@st.cache_data(show_spinner=False)
def _read_sample_df(file_path: str, delimiter: str, nrows: int, has_header: bool):
    if has_header:
        df = read_csv_robust(file_path, sep=delimiter, nrows=nrows, header=0)
        df = normalize_csv_headers_df(df)
        for c in df.columns:
            df[c] = df[c].astype(str).str.strip().replace({"nan": "", "None": ""})
        return df

    df = read_csv_robust(file_path, sep=delimiter, nrows=nrows, header=None)
    df.columns = [f"COL{i}" for i in range(df.shape[1])]
    df = normalize_csv_headers_df(df)
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip().replace({"nan": "", "None": ""})
    return df

@st.cache_data(show_spinner=False)
def infer_primary_keys_cached(file_path: str, delimiter: str, has_header: bool, sample_rows: int, max_pk_cols: int):
    df = _read_sample_df(file_path, delimiter, sample_rows, has_header)
    n = len(df)
    if n == 0:
        return ([], 0.0)

    cols = list(df.columns)

    if len(cols) == 1:
        c = cols[0]
        uniq = float(df[c].nunique(dropna=False) / n) if n else 0.0
        return ([c], uniq)

    cols2 = []
    for c in cols:
        if looks_like_measure_name(c) and not looks_like_id_name(c) and not looks_like_strong_id_name(c):
            continue
        cols2.append(c)

    cols3 = []
    for c in cols2:
        if looks_like_id_name(c) or looks_like_strong_id_name(c):
            cols3.append(c)
            continue
        num_ratio, dec_ratio = numeric_profile(df[c])
        if num_ratio >= 0.85 and dec_ratio >= 0.30:
            continue
        cols3.append(c)

    if not cols3:
        cols3 = cols[:]

    def score_col(c):
        s = df[c]
        non_empty = (s != "").mean()
        uniq_ratio = s.nunique(dropna=False) / n
        strong_bonus = 0.55 if looks_like_strong_id_name(c) else 0.0
        id_bonus = 0.25 if looks_like_id_name(c) else 0.0
        num_ratio, _ = numeric_profile(s)
        num_penalty = 0.15 if (not looks_like_id_name(c) and not looks_like_strong_id_name(c) and num_ratio >= 0.85) else 0.0
        return (uniq_ratio * 0.55) + (non_empty * 0.35) + strong_bonus + id_bonus - num_penalty

    ranked = sorted(cols3, key=score_col, reverse=True)
    strong_cols = [c for c in ranked if looks_like_strong_id_name(c)]
    non_strong = [c for c in ranked if c not in strong_cols]
    ranked = strong_cols + non_strong

    top = ranked[:TOP_COLS]

    def combo_uniqueness(combo):
        tmp = df[list(combo)].fillna("").astype(str).agg("||".join, axis=1)
        return float(tmp.nunique(dropna=False) / n)

    best_combo = []
    best_uniq = 0.0

    for k in range(1, min(max_pk_cols, len(top)) + 1):
        combos = list(combinations(top, k))
        if len(combos) > MAX_COMBOS_PER_K:
            combos = combos[:MAX_COMBOS_PER_K]

        improved = False
        for combo in combos:
            uniq = combo_uniqueness(combo)
            if uniq > best_uniq:
                best_uniq = uniq
                best_combo = list(combo)
                improved = True
            if uniq >= UNIQ_TARGET:
                combo_list = list(combo)
                for sc in strong_cols:
                    if sc not in combo_list and len(combo_list) < max_pk_cols:
                        combo_list.append(sc)
                forced_uniq = combo_uniqueness(tuple(combo_list)) if combo_list else uniq
                return (combo_list, forced_uniq)

        if k >= 3 and not improved:
            break

    if strong_cols:
        forced = list(dict.fromkeys(strong_cols + best_combo))
        forced = forced[:max_pk_cols]
        forced_uniq = combo_uniqueness(tuple(forced)) if forced else best_uniq
        return (forced, forced_uniq)

    return (best_combo, best_uniq)

def get_included_files():
    df_sel = st.session_state.get("df_sel")
    if df_sel is None or df_sel.empty:
        return []
    if "Include" not in df_sel.columns:
        return df_sel["File Name"].tolist()
    return df_sel[df_sel["Include"] == True]["File Name"].tolist()

# ----------------------------
# HEADER (con logo visible)
# ----------------------------
logo_uri = img_to_data_uri("images/logodefinitivo.png")

st.markdown(f"""
<div style="display:flex; align-items:center; gap:2px; transform: translateX(-115px);">
  <img src="{logo_uri}" style="height:300px; width:auto; display:block;" />
  <div class="mxl-title" style="margin:0; padding:0;
    font-family: Inter, Manrope, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
    font-size:132px; font-weight:600; letter-spacing:-2px; text-align:left;">
    <span style="color:#003264;">MX</span><span style="color:#058775;">Lab</span>
  </div>
</div>
<div class="mxl-subtitle">Internal EPAM FD Tool for MXtest Package Building, Execution and Results Reporting.</div>
<div class="mxl-divider"></div>
""", unsafe_allow_html=True)

def run_mxtest_package(pkg_dir_name: str, results_folder: str) -> subprocess.CompletedProcess:
    cmd_list = [
        "cmd.exe", "/c",
        "call", LAUNCH_CMD,
        pkg_dir_name,
        results_folder,
        FIXED_N
    ]
    return subprocess.run(
        cmd_list,
        cwd=MXTEST_ROOT,
        capture_output=True,
        text=True
    )

def _latest_run_folder(results_root: str) -> str | None:
    if not os.path.isdir(results_root):
        return None

    best_n = None
    best_path = None
    for name in os.listdir(results_root):
        m = re.match(r"^run_(\d+)$", str(name), flags=re.IGNORECASE)
        if not m:
            continue
        n = int(m.group(1))
        path = os.path.join(results_root, name)
        if not os.path.isdir(path):
            continue
        if best_n is None or n > best_n:
            best_n = n
            best_path = path
    return best_path

def find_output_xlsx_for_file(run_folder: str, file_name: str) -> str | None:
    if not run_folder or not os.path.isdir(run_folder):
        return None

    base = os.path.basename(file_name)
    stem = os.path.splitext(base)[0]

    tokens = {
        stem.lower(),
        (stem + "_csv").lower(),
        stem.replace(".", "_").lower(),
        (stem.replace(".", "_") + "_csv").lower(),
    }

    candidates = []
    for fn in os.listdir(run_folder):
        if not fn.lower().endswith(".xlsx"):
            continue
        low = fn.lower()
        if any(tok in low for tok in tokens):
            full = os.path.join(run_folder, fn)
            if os.path.isfile(full):
                candidates.append(full)

    if not candidates:
        return None

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]

def _read_statistics_from_xlsx(xlsx_path: str) -> dict:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = None
    for s in wb.sheetnames:
        if str(s).strip().lower() == "statistics":
            sheet = s
            break
    if sheet is None:
        wb.close()
        raise ValueError(f"Sheet 'Statistics' not found in {os.path.basename(xlsx_path)}")

    ws = wb[sheet]

    def v(cell):
        val = ws[cell].value
        return "" if val is None else val

    out = {
        "Number of row mismatches": v("C8"),
        "Number of row removed": v("C9"),
        "Number of row added": v("C10"),
        "Total number of rows (E)": v("C12"),
        "Total number of rows (R)": v("C13"),
    }
    wb.close()
    return out

def build_results_summary_df(run_folder: str, compared_files: list[str]) -> pd.DataFrame:
    rows = []
    for f in compared_files:
        out_xlsx = find_output_xlsx_for_file(run_folder, f)

        if not out_xlsx:
            rows.append({
                "File": f,
                "Output XLSX": "",
                "Status": "NOT COMPARED",
                "Number of row mismatches": "",
                "Number of row removed": "",
                "Number of row added": "",
                "Total number of rows (E)": "",
                "Total number of rows (R)": "",
                "Comparison Result": "N/A"
            })
            continue

        try:
            stats = _read_statistics_from_xlsx(out_xlsx)
            try:
                mismatches = int(stats.get("Number of row mismatches", 0) or 0)
                removed    = int(stats.get("Number of row removed", 0) or 0)
                added      = int(stats.get("Number of row added", 0) or 0)
            except Exception:
                mismatches = removed = added = 0

            comparison_result = (
                "Passed" if mismatches == 0 and removed == 0 and added == 0 else "Failed"
            )

            rows.append({
                "File": f,
                "Output XLSX": out_xlsx,
                "Status": "OK",
                "Number of row mismatches": stats.get("Number of row mismatches", ""),
                "Number of row removed": stats.get("Number of row removed", ""),
                "Number of row added": stats.get("Number of row added", ""),
                "Total number of rows (E)": stats.get("Total number of rows (E)", ""),
                "Total number of rows (R)": stats.get("Total number of rows (R)", ""),
                "Comparison Result": comparison_result
            })
        except Exception as e:
            rows.append({
                "File": f,
                "Output XLSX": out_xlsx,
                "Status": f"ERROR: {e}",
                "Number of row mismatches": "",
                "Number of row removed": "",
                "Number of row added": "",
                "Total number of rows (E)": "",
                "Total number of rows (R)": "",
                "Comparison Result": "ERROR"
            })

    return pd.DataFrame(rows)

def summary_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")

def summary_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Run Results Summary")
    return bio.getvalue()

# Session state defaults
if "deploy_ok" not in st.session_state:
    st.session_state.deploy_ok = None
if "deploy_path" not in st.session_state:
    st.session_state.deploy_path = None
if "deploy_err" not in st.session_state:
    st.session_state.deploy_err = None
if "pkg_dir" not in st.session_state:
    st.session_state.pkg_dir = None
if "last_run_ok" not in st.session_state:
    st.session_state.last_run_ok = None
if "last_run_out" not in st.session_state:
    st.session_state.last_run_out = None
if "last_run_err" not in st.session_state:
    st.session_state.last_run_err = None
if "last_run_cmd" not in st.session_state:
    st.session_state.last_run_cmd = None
if "last_run_folder" not in st.session_state:
    st.session_state.last_run_folder = None
if "results_summary_df" not in st.session_state:
    st.session_state.results_summary_df = None

# Status box
msg_box = st.container()
if st.session_state.deploy_ok is True and st.session_state.deploy_path:
    msg_box.success(f"Package deployed successfully to:\n{st.session_state.deploy_path}")
elif st.session_state.deploy_ok is False and st.session_state.deploy_err:
    msg_box.error(st.session_state.deploy_err)

# ----------------------------
# TABS
# ----------------------------
tab_main, tab_summary = st.tabs(["Main", "Run Results Summary"])
# =========================================================
# MAIN TAB
# =========================================================
with tab_main:
    col1, col2 = st.columns([2, 1])
    source_path = ""
    target_path = ""

    # LEFT PANEL
    with col1:
        with st.container(border=True):
            st.subheader("Package Configuration")
            p1, p2 = st.columns(2)
            with p1:
                package_name = st.text_input("Package Name")
            with p2:
                version = st.text_input("Package Version")

            st.subheader("Global Test Configuration")
            if "global_vars" not in st.session_state:
                st.session_state.global_vars = pd.DataFrame(columns=["Name", "Value"])

            global_vars_df = st.data_editor(
                st.session_state.global_vars,
                num_rows="dynamic",
                use_container_width=True
            )
            st.session_state.df_global_vars = global_vars_df
            st.session_state.global_vars = global_vars_df

            st.subheader("File Comparison Configuration")
            path1, path2 = st.columns(2)
            with path1:
                source_path = st.text_input("Source Files Path")
            with path2:
                target_path = st.text_input("Target Files Path")

            if "df_sel" not in st.session_state:
                st.session_state.df_sel = pd.DataFrame(columns=["File Name", "Include"])
            if "df_delimiters" not in st.session_state:
                st.session_state.df_delimiters = pd.DataFrame(columns=["File Name", "Delimiter"])
            if "df_fields" not in st.session_state:
                st.session_state.df_fields = pd.DataFrame(columns=["File Name", "Has Header", "Fields"])
            if "df_pks" not in st.session_state:
                st.session_state.df_pks = pd.DataFrame(columns=["File Name", "Suggested PKs", "Uniqueness", "Use PK", "Primary Keys"])

            selected_files = []

            if source_path and target_path and os.path.exists(source_path) and os.path.exists(target_path):
                ALLOWED_EXT = {".csv", ".txt", ".dat"}

                def list_only_files(folder):
                    out = []
                    for name in os.listdir(folder):
                        full = os.path.join(folder, name)
                        if os.path.isfile(full):
                            ext = os.path.splitext(name)[1].lower()
                            if (not ALLOWED_EXT) or (ext in ALLOWED_EXT):
                                out.append(name)
                    return set(out)

                source_files = list_only_files(source_path)
                target_files = list_only_files(target_path)
                matched_files = sorted(list(source_files & target_files))

                if matched_files:
                    df_files = pd.DataFrame({"File Name": matched_files, "Include": [True] * len(matched_files)})

                    st.markdown("#### Matched Files (select which to include)")
                    df_sel = st.data_editor(df_files, num_rows="dynamic", use_container_width=True, key="file_selection")
                    st.session_state.df_sel = df_sel

                    if "Include" not in df_sel.columns:
                        df_sel["Include"] = True

                    selected_files = df_sel[df_sel["Include"] == True]["File Name"].tolist()

                    st.markdown("#### File Delimiters")
                    delimiter_rows = [{"File Name": f, "Delimiter": detect_delimiter(os.path.join(source_path, f))} for f in selected_files]
                    df_delimiters = pd.DataFrame(delimiter_rows)
                    df_delimiters = st.data_editor(df_delimiters, use_container_width=True, key="file_delimiters")
                    st.session_state.df_delimiters = df_delimiters

                    st.markdown("#### File Fields")
                    fields_rows = []
                    for f in selected_files:
                        delimiter = df_delimiters[df_delimiters["File Name"] == f]["Delimiter"].values[0]
                        path = os.path.join(source_path, f)
                        try:
                            with open(path, "r", encoding="utf-8", errors="replace") as file:
                                first_line = file.readline().strip()
                            raw_fields = first_line.split(delimiter)

                            numeric_like = sum(x.strip().replace(".", "", 1).isdigit() for x in raw_fields)
                            empty_like = sum(x.strip() == "" for x in raw_fields)
                            not_header = (numeric_like + empty_like) >= max(1, int(0.6 * len(raw_fields)))
                            has_header = not not_header

                            if has_header:
                                fields_norm = normalize_headers_list([x.strip() for x in raw_fields])
                                fields_str = ", ".join(fields_norm)
                            else:
                                fields_norm = [f"COL{i}" for i in range(len(raw_fields))]
                                fields_str = ", ".join(fields_norm)
                        except Exception:
                            has_header = True
                            fields_str = ""

                        fields_rows.append({"File Name": f, "Has Header": has_header, "Fields": fields_str})

                    df_fields = pd.DataFrame(fields_rows)
                    df_fields = st.data_editor(
                        df_fields,
                        use_container_width=True,
                        key="file_fields",
                        column_config={
                            "Has Header": st.column_config.CheckboxColumn(
                                "Has Header",
                                help="Uncheck if the file has no header row (COL0..COLN will be assumed)."
                            )
                        }
                    )

                    if not df_fields.empty and "Fields" in df_fields.columns:
                        df_fields["Fields"] = df_fields["Fields"].apply(
                            lambda s: ", ".join(normalize_headers_list([x.strip() for x in str(s).split(",")]))
                        )

                    st.session_state.df_fields = df_fields

    # RIGHT PANEL
    with col2:
        with st.container(border=True):
            st.subheader("Build Control")
            st.caption("Generate MXtest Package")
            build_button = st.button("🚀 Build MXtest Package", use_container_width=True)

            if st.session_state.deploy_ok is True and st.session_state.deploy_path:
                st.success(f"Deployed to:\n{st.session_state.deploy_path}")
            elif st.session_state.deploy_ok is False and st.session_state.deploy_err:
                st.error(st.session_state.deploy_err)

            # Run Control (only after build OK)
            if st.session_state.deploy_ok is True and st.session_state.pkg_dir:
                st.divider()
                st.subheader("Run Control")

                results_folder = st.text_input("Results folder", value=DEFAULT_RESULTS_FOLDER)
                run_button = st.button("▶ Run MXtest Package", use_container_width=True)

                if run_button:
                    st.session_state.last_run_ok = None
                    st.session_state.last_run_out = None
                    st.session_state.last_run_err = None
                    st.session_state.last_run_cmd = None
                    st.session_state.last_run_folder = None
                    st.session_state.results_summary_df = None

                    pkg_dir = st.session_state.pkg_dir
                    st.session_state.last_run_cmd = (
                        f'cmd.exe /c call "{LAUNCH_CMD}" "{pkg_dir}" "{results_folder}" {FIXED_N}'
                    )

                    with st.spinner("Running package..."):
                        run_res = run_mxtest_package(pkg_dir, results_folder)

                    st.session_state.last_run_ok = (run_res.returncode == 0)
                    st.session_state.last_run_out = run_res.stdout
                    st.session_state.last_run_err = run_res.stderr

                    if run_res.returncode == 0:
                        st.success("Run finished OK.")

                        run_folder = _latest_run_folder(FC_RESULTS_ROOT)
                        st.session_state.last_run_folder = run_folder

                        compared_files = get_included_files()
                        if run_folder and compared_files:
                            with st.spinner("Building results summary..."):
                                st.session_state.results_summary_df = build_results_summary_df(run_folder, compared_files)

                    else:
                        st.error(f"Run failed (code={run_res.returncode}).")

                if st.session_state.last_run_cmd or st.session_state.last_run_out or st.session_state.last_run_err:
                    with st.expander("Run Debug Output", expanded=False):
                        if st.session_state.last_run_cmd:
                            st.code("COMMAND:\n" + st.session_state.last_run_cmd)
                        st.code("STDOUT:\n" + (st.session_state.last_run_out or "(empty)"))
                        st.code("STDERR:\n" + (st.session_state.last_run_err or "(empty)"))

    # PK section (unchanged)
    included_files = get_included_files()

    if source_path and target_path and os.path.exists(source_path) and os.path.exists(target_path) and included_files:
        st.markdown("### Primary Key Suggestions")
        st.caption(f"Fast mode: samples {SAMPLE_ROWS} rows. Infers max {MAX_FILES_INFERENCE} files per click.")

        infer_btn = st.button("🔎 Infer PKs (fast)", use_container_width=False)

        if infer_btn:
            included_files = get_included_files()
            files_for_inference = included_files[:MAX_FILES_INFERENCE]

            if len(included_files) > MAX_FILES_INFERENCE:
                st.warning(
                    f"Too many files selected ({len(included_files)}). "
                    f"Inferring PKs only for first {MAX_FILES_INFERENCE}. "
                    f"Run again after narrowing selection."
                )

            pk_rows = []
            with st.spinner("Inferring PKs..."):
                for f in files_for_inference:
                    file_path = os.path.join(source_path, f)
                    if not os.path.isfile(file_path):
                        continue

                    df_delimiters = st.session_state.df_delimiters
                    delimiter = ","
                    if df_delimiters is not None and not df_delimiters.empty:
                        v = df_delimiters[df_delimiters["File Name"] == f]["Delimiter"]
                        if len(v) > 0:
                            delimiter = v.values[0]

                    df_fields = st.session_state.df_fields
                    has_header = True
                    if df_fields is not None and not df_fields.empty and "Has Header" in df_fields.columns:
                        vv = df_fields.loc[df_fields["File Name"] == f, "Has Header"]
                        if len(vv) > 0:
                            has_header = str(vv.values[0]).strip().lower() in ("true", "1", "yes", "y")

                    keys, uniq = infer_primary_keys_cached(file_path, delimiter, has_header, SAMPLE_ROWS, MAX_PK_COLS)
                    keys = normalize_pk_list(keys)

                    pk_rows.append({
                        "File Name": f,
                        "Suggested PKs": ", ".join(keys),
                        "Uniqueness": round(uniq * 100, 2),
                        "Use PK": False,
                        "Primary Keys": ", ".join(keys)
                    })

            df_pks_new = pd.DataFrame(pk_rows)

            if not st.session_state.df_pks.empty and "File Name" in st.session_state.df_pks.columns:
                old = st.session_state.df_pks.set_index("File Name")
                df_pks_new = df_pks_new.set_index("File Name")
                for idx in df_pks_new.index:
                    if idx in old.index:
                        df_pks_new.loc[idx, "Use PK"] = old.loc[idx, "Use PK"]
                        df_pks_new.loc[idx, "Primary Keys"] = old.loc[idx, "Primary Keys"]
                df_pks_new = df_pks_new.reset_index()

                kept = st.session_state.df_pks[~st.session_state.df_pks["File Name"].isin(files_for_inference)]
                st.session_state.df_pks = pd.concat([kept, df_pks_new], ignore_index=True)
            else:
                st.session_state.df_pks = df_pks_new

        b1, b2 = st.columns([1, 1])
        with b1:
            if st.button("✅ Use PK en TODOS", use_container_width=True, disabled=st.session_state.df_pks.empty):
                if not st.session_state.df_pks.empty and "Use PK" in st.session_state.df_pks.columns:
                    st.session_state.df_pks["Use PK"] = True
                    st.rerun()
        with b2:
            if st.button("❌ Quitar Use PK en TODOS", use_container_width=True, disabled=st.session_state.df_pks.empty):
                if not st.session_state.df_pks.empty and "Use PK" in st.session_state.df_pks.columns:
                    st.session_state.df_pks["Use PK"] = False
                    st.rerun()

        df_pks = st.data_editor(
            st.session_state.df_pks,
            use_container_width=True,
            key="pk_suggestions_editor",
            column_config={
                "File Name": st.column_config.TextColumn(width="medium"),
                "Suggested PKs": st.column_config.TextColumn(width="large"),
                "Uniqueness": st.column_config.NumberColumn(width="small"),
                "Use PK": st.column_config.CheckboxColumn(width="small"),
                "Primary Keys": st.column_config.TextColumn(width="large"),
            },
        )

        if df_pks is not None and not df_pks.empty:
            if "Primary Keys" in df_pks.columns:
                df_pks["Primary Keys"] = df_pks["Primary Keys"].apply(
                    lambda s: ", ".join(normalize_pk_list([x.strip() for x in str(s).split(",")]))
                )
            if "Suggested PKs" in df_pks.columns:
                df_pks["Suggested PKs"] = df_pks["Suggested PKs"].apply(
                    lambda s: ", ".join(normalize_pk_list([x.strip() for x in str(s).split(",")]))
                )

        st.session_state.df_pks = df_pks
# =========================================================
# SUMMARY TAB
# =========================================================
with tab_summary:
    st.markdown("### Run Results Summary")

    if st.session_state.last_run_ok is True and isinstance(st.session_state.results_summary_df, pd.DataFrame):
        pkg_dir_val = st.session_state.get("pkg_dir") or ""
        pkg_name_val = ""
        pkg_ver_val = ""
        if "_" in pkg_dir_val:
            pkg_name_val = pkg_dir_val.rsplit("_", 1)[0]
            pkg_ver_val = pkg_dir_val.rsplit("_", 1)[1]

        left, right = st.columns([2, 1])
        with left:
            st.markdown("#### Package Info")
            st.json({
                "Package Name": pkg_name_val,
                "Version": pkg_ver_val,
                "Compared files": ", ".join(get_included_files())
            })

        with right:
            st.markdown("#### Output Location")
            st.write("Latest run folder:")
            st.code(st.session_state.last_run_folder or "(not found)")

        df_export = st.session_state.results_summary_df.copy()
        df_show = df_export.copy()

        if "Output XLSX" in df_show.columns:
            df_show["Output XLSX"] = df_show["Output XLSX"].apply(
                lambda p: os.path.basename(p) if isinstance(p, str) else p
            )

        st.dataframe(df_show, use_container_width=True)

        st.markdown("#### Export")
        c1, c2 = st.columns(2)

        with c1:
            with st.container():
                st.markdown('<div class="custom-download-btn">', unsafe_allow_html=True)
                st.download_button(
                    label="⬇️ Download summary (CSV)",
                    data=summary_to_csv_bytes(df_export),
                    file_name=f"RunResultsSummary_{pkg_dir_val or 'package'}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                st.markdown('</div>', unsafe_allow_html=True)

        with c2:
            with st.container():
                st.markdown('<div class="custom-download-btn">', unsafe_allow_html=True)
                st.download_button(
                    label="⬇️ Download summary (Excel)",
                    data=summary_to_xlsx_bytes(df_export),
                    file_name=f"RunResultsSummary_{pkg_dir_val or 'package'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.info("No run results to show yet. Execute a package successfully first.")
# =========================================================
# BUILD PROCESS
# =========================================================
if 'build_button' in locals() and build_button:
    st.session_state.deploy_ok = None
    st.session_state.deploy_path = None
    st.session_state.deploy_err = None

    st.session_state.pkg_dir = None
    st.session_state.last_run_ok = None
    st.session_state.last_run_out = None
    st.session_state.last_run_err = None
    st.session_state.last_run_cmd = None
    st.session_state.last_run_folder = None
    st.session_state.results_summary_df = None

    with st.spinner("Building package... please wait"):
        if package_name.strip() == "":
            st.session_state.deploy_ok = False
            st.session_state.deploy_err = "Package Name is required"
            st.rerun()
        if version.strip() == "":
            st.session_state.deploy_ok = False
            st.session_state.deploy_err = "Package Version is required"
            st.rerun()

        if not source_path or not target_path:
            st.session_state.deploy_ok = False
            st.session_state.deploy_err = "Source Files Path and Target Files Path are required."
            st.rerun()

        if not (os.path.exists(source_path) and os.path.exists(target_path)):
            st.session_state.deploy_ok = False
            st.session_state.deploy_err = "Source/Target path does not exist."
            st.rerun()

        if st.session_state.df_sel is None or st.session_state.df_sel.empty:
            st.session_state.deploy_ok = False
            st.session_state.deploy_err = "No matched files selected. Please provide valid paths and select at least one file."
            st.rerun()

        config = {
            "package": {"name": package_name, "version": version},
            "test_suite": "File comparisons",
            "global_variables": {},
            "file_comparison": {"source_path": source_path, "target_path": target_path},
            "files": []
        }

        for _, row in st.session_state.global_vars.iterrows():
            if str(row.get("Name", "")).strip() != "":
                config["global_variables"][row["Name"]] = row.get("Value", "")

        df_sel = st.session_state.df_sel
        df_delimiters = st.session_state.df_delimiters
        df_fields = st.session_state.df_fields
        df_pks = st.session_state.df_pks if "df_pks" in st.session_state else pd.DataFrame()

        if "Include" not in df_sel.columns:
            df_sel["Include"] = True
        selected_files_build = df_sel[df_sel["Include"] == True]["File Name"].tolist()

        for f in selected_files_build:
            delimiter = df_delimiters[df_delimiters["File Name"] == f]["Delimiter"].values[0]

            fields_str = df_fields[df_fields["File Name"] == f]["Fields"].values[0] if not df_fields.empty else ""
            fields = [x.strip() for x in str(fields_str).split(",") if x.strip()]
            fields = normalize_headers_list(fields)

            primary_keys = []
            use_pk = False
            if not df_pks.empty:
                rowpk = df_pks[df_pks["File Name"] == f]
                if not rowpk.empty:
                    use_pk = bool(rowpk["Use PK"].values[0])
                    if use_pk:
                        pk_str = str(rowpk["Primary Keys"].values[0])
                        primary_keys = [x.strip() for x in pk_str.split(",") if x.strip()]
                        primary_keys = normalize_pk_list(primary_keys)

            config["files"].append({
                "name": f,
                "delimiter": delimiter,
                "fields": fields,
                "use_primary_keys": use_pk,
                "primary_keys": primary_keys
            })

        with tempfile.NamedTemporaryFile(delete=False, suffix=".json", mode="w", encoding="utf-8") as tmp:
            json.dump(config, tmp, indent=4)
            config_path = tmp.name

        command = ["python", "build_mxtest.py", config_path]
        result = subprocess.run(command, capture_output=True, text=True)

    with st.expander("Build Output", expanded=False):
        st.code(result.stdout)
        if result.returncode != 0:
            st.code(result.stderr)

    err_txt = (result.stderr or "").strip()
    out_txt = (result.stdout or "").strip()

    if result.returncode != 0:
        st.session_state.deploy_ok = False
        st.session_state.deploy_err = (
            "Build failed.\n\n"
            "---- STDERR ----\n" + (err_txt if err_txt else "(empty)") + "\n\n"
            "---- STDOUT ----\n" + (out_txt if out_txt else "(empty)")
        )
        st.session_state.deploy_path = None
        st.session_state.pkg_dir = None
        st.rerun()

    pkg_dir = f"{package_name}_{version}"
    try:
        final_path = move_package_to_final_location(pkg_dir, FINAL_OUTPUT_DIR)
        st.session_state.deploy_ok = True
        st.session_state.deploy_path = final_path
        st.session_state.deploy_err = None
        st.session_state.pkg_dir = pkg_dir
    except Exception as e:
        st.session_state.deploy_ok = False
        st.session_state.deploy_err = str(e)
        st.session_state.deploy_path = None
        st.session_state.pkg_dir = None

    st.rerun()

st.markdown('<div class="author-footer">Developed by Juan Alberto Villagrán Guerrero</div>', unsafe_allow_html=True)